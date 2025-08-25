# streamlit_excel_compare.py
# -*- coding: utf-8 -*-
import io
import math
import tempfile
from typing import Optional, Tuple, List, Dict, Any

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter


# ----------------------------
# Utilities
# ----------------------------
def normalize_colnames(cols: List[str]) -> List[str]:
    out = []
    for c in cols:
        s = str(c).strip()
        s = " ".join(s.split())
        out.append(s)
    return out


def to_str(x) -> str:
    if pd.isna(x):
        return ""
    return str(x)


def normalize_series(s: pd.Series, strip_whitespace: bool = True, case_sensitive: bool = False) -> pd.Series:
    ss = s.astype("object").map(to_str)
    if strip_whitespace:
        ss = ss.map(lambda v: v.strip())
        ss = ss.map(lambda v: " ".join(v.split()))
    if not case_sensitive:
        ss = ss.map(lambda v: v.lower())
    return ss


def try_parse_float(val: str, decimal_comma: bool = False) -> Optional[float]:
    if val is None or val == "" or pd.isna(val):
        return None
    s = str(val).strip()
    if s == "":
        return None
    if decimal_comma:
        if s.count(",") == 1 and s.count(".") == 0:
            s = s.replace(",", ".")
        s = s.replace("\u202f", "").replace(" ", "")
    try:
        return float(s)
    except Exception:
        return None


def equal_values(a, b,
                 strip_whitespace: bool,
                 case_sensitive: bool,
                 numeric_tol: float,
                 decimal_comma: bool) -> bool:
    if (a is None or a == "" or (isinstance(a, float) and math.isnan(a)) or pd.isna(a)) and \
       (b is None or b == "" or (isinstance(b, float) and math.isnan(b)) or pd.isna(b)):
        return True

    if numeric_tol is not None and numeric_tol > 0:
        fa = try_parse_float(a, decimal_comma=decimal_comma)
        fb = try_parse_float(b, decimal_comma=decimal_comma)
        if fa is not None and fb is not None:
            return abs(fa - fb) <= numeric_tol

    sa = to_str(a)
    sb = to_str(b)
    if strip_whitespace:
        sa = " ".join(sa.strip().split())
        sb = " ".join(sb.strip().split())
    if not case_sensitive:
        sa = sa.lower()
        sb = sb.lower()
    return sa == sb


def choose_best_key(dfA: pd.DataFrame, dfB: pd.DataFrame) -> Tuple[str, pd.Series, pd.Series, Dict[str, Dict[str, int]]]:
    common_cols = [c for c in dfA.columns if c in dfB.columns]
    if not common_cols:
        raise ValueError("Aucune colonne commune entre les deux fichiers.")

    metrics = {}
    best = None
    best_overlap = -1
    best_minfilled = -1

    for c in common_cols:
        sA = dfA[c]
        sB = dfB[c]
        nonemptyA = sA.notna() & (sA.astype(str).str.strip() != "")
        nonemptyB = sB.notna() & (sB.astype(str).str.strip() != "")
        countA = int(nonemptyA.sum())
        countB = int(nonemptyB.sum())

        normA = normalize_series(sA, strip_whitespace=True, case_sensitive=False)[nonemptyA]
        normB = normalize_series(sB, strip_whitespace=True, case_sensitive=False)[nonemptyB]
        setA = set(normA.tolist())
        setB = set(normB.tolist())
        overlap = len(setA & setB)
        minfilled = min(countA, countB)

        metrics[c] = {"filled_in_A": countA, "filled_in_B": countB, "overlap_values": overlap, "min_filled": minfilled}

        if overlap > best_overlap or (overlap == best_overlap and minfilled > best_minfilled):
            best = c
            best_overlap = overlap
            best_minfilled = minfilled

    key = best
    keyA = normalize_series(dfA[key], strip_whitespace=True, case_sensitive=False)
    keyB = normalize_series(dfB[key], strip_whitespace=True, case_sensitive=False)
    return key, keyA, keyB, metrics


def list_sheet_names(xls_bytes: bytes) -> List[str]:
    with pd.ExcelFile(io.BytesIO(xls_bytes)) as xf:
        return xf.sheet_names


def read_df_from_bytes(xls_bytes: bytes, sheet_name: Any) -> pd.DataFrame:
    df = pd.read_excel(io.BytesIO(xls_bytes), sheet_name=sheet_name)
    if isinstance(df, dict):
        first_name = next(iter(df))
        df = df[first_name]
    return df


def write_sorted_sheet(ws, df: pd.DataFrame, title: str):
    # Write header
    header_font = Font(bold=True)
    ws.title = title[:31]  # Excel sheet name limit
    for j, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=j, value=str(col))
        cell.font = header_font
    # Write data
    for i in range(len(df)):
        row_vals = df.iloc[i].tolist()
        for j, val in enumerate(row_vals, start=1):
            ws.cell(row=i+2, column=j, value=val)
    ws.freeze_panes = "A2"
    # Auto filter
    ws.auto_filter.ref = ws.dimensions
    # Simple width heuristic
    for j, col in enumerate(df.columns, start=1):
        width = min(max(10, int(df[col].astype(str).map(len).max() if len(df) else 10) + 2), 60)
        ws.column_dimensions[get_column_letter(j)].width = width


def compare_single_sheet(sheet: str,
                         bytes_a: bytes,
                         bytes_b: bytes,
                         numeric_tol: float,
                         decimal_comma: bool,
                         case_sensitive: bool,
                         strip_ws: bool):
    dfA = read_df_from_bytes(bytes_a, sheet)
    dfB = read_df_from_bytes(bytes_b, sheet)

    dfA.columns = normalize_colnames(dfA.columns.tolist())
    dfB.columns = normalize_colnames(dfB.columns.tolist())

    common_cols = [c for c in dfA.columns if c in dfB.columns]
    if not common_cols:
        return None  # cannot compare this sheet

    key, keyA_norm, keyB_norm, key_metrics = choose_best_key(dfA, dfB)

    dfA2 = dfA.copy(); dfB2 = dfB.copy()
    dfA2["_key_norm_"] = keyA_norm
    dfB2["_key_norm_"] = keyB_norm
    dfA2["_orig_pos_"] = np.arange(len(dfA2))
    dfB2["_orig_pos_"] = np.arange(len(dfB2))

    dfA_sorted = dfA2.sort_values(by=["_key_norm_", "_orig_pos_"], kind="mergesort").reset_index(drop=True)
    dfB_sorted = dfB2.sort_values(by=["_key_norm_", "_orig_pos_"], kind="mergesort").reset_index(drop=True)

    max_len = max(len(dfA_sorted), len(dfB_sorted))
    dfA_cmp = dfA_sorted.reindex(range(max_len)).reset_index(drop=True)
    dfB_cmp = dfB_sorted.reindex(range(max_len)).reset_index(drop=True)

    common_cols_no_aux = [c for c in common_cols if c not in ["_key_norm_", "_orig_pos_"]]

    # Build differences
    diff_records = []
    total_cells = 0
    diff_cells = 0

    def get_val(df, i, col):
        try:
            return df.at[i, col]
        except Exception:
            return np.nan

    for i in range(max_len):
        key_val = get_val(dfA_cmp, i, key)
        if pd.isna(key_val) or str(key_val).strip() == "":
            key_val = get_val(dfB_cmp, i, key)
        for col in common_cols_no_aux:
            a = get_val(dfA_cmp, i, col)
            b = get_val(dfB_cmp, i, col)
            eq = equal_values(a, b, strip_ws, case_sensitive, numeric_tol, decimal_comma)
            total_cells += 1
            if not eq:
                diff_cells += 1
                diff_records.append({
                    "sheet": sheet,
                    "row_index_sorted": i + 1,
                    "key_value": to_str(key_val),
                    "column": col,
                    "value_in_A": a,
                    "value_in_B": b
                })

    differences_df = pd.DataFrame(diff_records, columns=[
        "sheet", "row_index_sorted", "key_value", "column", "value_in_A", "value_in_B"
    ])

    # Keys stats
    keysA = dfA_sorted["_key_norm_"][dfA_sorted["_key_norm_"] != ""].astype(str)
    keysB = dfB_sorted["_key_norm_"][dfB_sorted["_key_norm_"] != ""].astype(str)
    setA = set(keysA.tolist())
    setB = set(keysB.tolist())
    only_in_A = sorted(list(setA - setB))
    only_in_B = sorted(list(setB - setA))

    dupA = keysA[keysA.duplicated(keep=False)].value_counts().reset_index()
    dupA.columns = ["key_value", "count_in_A"]
    dupB = keysB[keysB.duplicated(keep=False)].value_counts().reset_index()
    dupB.columns = ["key_value", "count_in_B"]
    duplicate_keys = pd.merge(dupA, dupB, on="key_value", how="outer").fillna(0)
    duplicate_keys["count_in_A"] = duplicate_keys["count_in_A"].astype(int)
    duplicate_keys["count_in_B"] = duplicate_keys["count_in_B"].astype(int)
    duplicate_keys = duplicate_keys.sort_values(by=["count_in_A", "count_in_B", "key_value"], ascending=[False, False, True])

    # Sheet-level summary
    summary = {
        "sheet": sheet,
        "chosen_key": key,
        "rows_in_A": int(len(dfA_sorted)),
        "rows_in_B": int(len(dfB_sorted)),
        "common_columns_compared": ", ".join(common_cols_no_aux),
        "total_cells_compared": int(total_cells),
        "different_cells": int(diff_cells),
        "different_cells_pct": float((diff_cells/total_cells*100) if total_cells else 0.0),
        "keys_only_in_A": int(len(only_in_A)),
        "keys_only_in_B": int(len(only_in_B)),
        "duplicate_key_values": int(duplicate_keys.shape[0]),
    }

    return {
        "sheet": sheet,
        "summary": summary,
        "dfA_sorted": dfA_sorted.drop(columns=["_key_norm_", "_orig_pos_"]),
        "dfB_sorted": dfB_sorted.drop(columns=["_key_norm_", "_orig_pos_"]),
        "common_cols": common_cols_no_aux,
        "differences": differences_df,
        "only_in_A": pd.DataFrame({"sheet":[sheet]*len(only_in_A), "missing_in_B_key": only_in_A}),
        "only_in_B": pd.DataFrame({"sheet":[sheet]*len(only_in_B), "missing_in_A_key": only_in_B}),
        "duplicate_keys": duplicate_keys.assign(sheet=sheet),
    }


def build_report_for_all_sheets(bytes_a: bytes,
                                bytes_b: bytes,
                                numeric_tol: float,
                                decimal_comma: bool,
                                case_sensitive: bool,
                                strip_ws: bool):
    sheets_a = set(list_sheet_names(bytes_a))
    sheets_b = set(list_sheet_names(bytes_b))
    common_sheets = [s for s in sheets_a if s in sheets_b]
    if not common_sheets:
        raise ValueError("Aucune feuille commune entre A et B.")

    results = []
    for s in common_sheets:
        r = compare_single_sheet(s, bytes_a, bytes_b, numeric_tol, decimal_comma, case_sensitive, strip_ws)
        if r is not None:
            results.append(r)

    if not results:
        raise ValueError("Aucune feuille comparable (pas de colonnes communes).")

    # Aggregate analytics
    summary_rows = [r["summary"] for r in results]
    summary_df = pd.DataFrame(summary_rows)

    differences_df = pd.concat([r["differences"] for r in results], ignore_index=True)
    only_in_A_df = pd.concat([r["only_in_A"] for r in results], ignore_index=True)
    only_in_B_df = pd.concat([r["only_in_B"] for r in results], ignore_index=True)
    duplicate_keys_df = pd.concat([r["duplicate_keys"] for r in results], ignore_index=True)

    # Create workbook with sorted sheets for A and B, highlighting diffs
    wb = Workbook()
    # Remove default sheet
    default_ws = wb.active
    wb.remove(default_ws)

    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

    for r in results:
        sheet = r["sheet"]
        # A sorted
        wsA = wb.create_sheet(title=(sheet[:25] + "_A_sorted"))
        write_sorted_sheet(wsA, r["dfA_sorted"], wsA.title)
        # B sorted
        wsB = wb.create_sheet(title=(sheet[:25] + "_B_sorted"))
        write_sorted_sheet(wsB, r["dfB_sorted"], wsB.title)

        # Maps for quick column index lookup (normalize header names)
        def header_map(ws):
            m = {}
            max_col = ws.max_column
            for c in range(1, max_col+1):
                name = ws.cell(row=1, column=c).value
                if name is None:
                    continue
                key = " ".join(str(name).strip().split()).lower()
                if key not in m:
                    m[key] = c
            return m

        mapA = header_map(wsA)
        mapB = header_map(wsB)

        # Apply highlights based on differences after sorting
        for _, diff in r["differences"].iterrows():
            col_norm = " ".join(str(diff["column"]).strip().split()).lower()
            row_sorted = int(diff["row_index_sorted"])
            # highlight only if column exists in that sheet (should be true)
            if col_norm in mapA:
                wsA.cell(row=1 + row_sorted, column=mapA[col_norm]).fill = red_fill
            if col_norm in mapB:
                wsB.cell(row=1 + row_sorted, column=mapB[col_norm]).fill = red_fill

    # Save to temp and append analytics
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name
    wb.save(tmp_path)

    with pd.ExcelWriter(tmp_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        summary_df.to_excel(writer, index=False, sheet_name="Summary_by_sheet")
        differences_df.to_excel(writer, index=False, sheet_name="Differences_all")
        only_in_A_df.to_excel(writer, index=False, sheet_name="Only_in_A")
        only_in_B_df.to_excel(writer, index=False, sheet_name="Only_in_B")
        duplicate_keys_df.to_excel(writer, index=False, sheet_name="Duplicate_keys")

    with open(tmp_path, "rb") as f:
        out_bytes = f.read()

    return out_bytes, summary_df.head(100), differences_df.head(300)


# ----------------------------
# Streamlit UI
# ----------------------------
st.set_page_config(page_title="Comparateur Excel Quality", layout="wide")
st.title("Comparateur Excel")

st.markdown("""
Dépose deux classeurs **.xlsx** (A & B). Pour **chaque feuille commune** :
1. détection de la **meilleure colonne-clé** commune,
2. **tri** des deux feuilles sur cette clé,
3. **comparaison cellule par cellule** des **colonnes communes**,
4. génération de deux feuilles **_A_sorted** et **_B_sorted** avec **les cellules différentes en rouge**,
5. ajout d'onglets analytiques (Summary_by_sheet, Differences_all, Only_in_A, Only_in_B, Duplicate_keys).
""")

with st.sidebar:
    st.header("⚙️ Options")
    numeric_tol = st.number_input("Tolérance numérique", min_value=0.0, value=0.0, step=0.0001, format="%.6f")
    decimal_comma = st.checkbox("Virgule décimale (européenne)", value=False)
    case_sensitive = st.checkbox("Sensible à la casse", value=False)
    strip_ws = st.checkbox("Nettoyer espaces superflus", value=True)
    output_name = st.text_input("Nom du fichier de sortie", value="Differences.xlsx")

col1, col2 = st.columns(2)
with col1:
    up_a = st.file_uploader("📄 Classeur A (.xlsx)", type=["xlsx"], accept_multiple_files=False)
with col2:
    up_b = st.file_uploader("📄 Classeur B (.xlsx)", type=["xlsx"], accept_multiple_files=False)

run = st.button("Générer le rapport multi-feuilles")

if run:
    if up_a is None or up_b is None:
        st.warning("Merci de charger **A** et **B**.")
    else:
        bytes_a = up_a.getvalue()
        bytes_b = up_b.getvalue()
        try:
            with st.spinner("Comparaison multi-feuilles en cours..."):
                out_bytes, preview_summary, preview_diffs = build_report_for_all_sheets(
                    bytes_a, bytes_b, numeric_tol, decimal_comma, case_sensitive, strip_ws
                )
            st.success("Rapport généré.")

            st.subheader("Aperçu — Summary_by_sheet (top 100)")
            st.dataframe(preview_summary, use_container_width=True)

            st.subheader("Aperçu — Differences_all (top 300)")
            st.dataframe(preview_diffs, use_container_width=True)

            st.download_button(
                label="Télécharger le rapport Excel",
                data=out_bytes,
                file_name=output_name or "Differences.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as e:
            st.error(f"Erreur : {e}")
            st.exception(e)
