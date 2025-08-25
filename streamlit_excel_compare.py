# streamlit_excel_compare.py
# -*- coding: utf-8 -*-
import io, math, tempfile
from typing import List, Any
import numpy as np, pandas as pd, streamlit as st
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

def normalize_colnames(cols: List[str]) -> List[str]:
    return [' '.join(str(c).strip().split()) for c in cols]

def to_str(x): return '' if pd.isna(x) else str(x)

def normalize_series(s: pd.Series, strip_whitespace=True, case_sensitive=False) -> pd.Series:
    ss = s.astype('object').map(to_str)
    if strip_whitespace: ss = ss.map(lambda v: ' '.join(v.strip().split()))
    if not case_sensitive: ss = ss.map(lambda v: v.lower())
    return ss

def try_parse_float(val, decimal_comma=False):
    if val is None or val == '' or pd.isna(val): return None
    s = str(val).strip()
    if s=='': return None
    if decimal_comma:
        if s.count(',')==1 and s.count('.')==0: s = s.replace(',', '.')
        s = s.replace('\u202f','').replace(' ','')
    try: return float(s)
    except: return None

def equal_values(a,b,strip_whitespace,case_sensitive,numeric_tol,decimal_comma):
    if (a in [None,''] or (isinstance(a,float) and math.isnan(a)) or pd.isna(a)) and        (b in [None,''] or (isinstance(b,float) and math.isnan(b)) or pd.isna(b)): return True
    if numeric_tol and numeric_tol>0:
        fa, fb = try_parse_float(a,decimal_comma), try_parse_float(b,decimal_comma)
        if fa is not None and fb is not None: return abs(fa-fb) <= numeric_tol
    sa, sb = to_str(a), to_str(b)
    if strip_whitespace: sa, sb = ' '.join(sa.strip().split()), ' '.join(sb.strip().split())
    if not case_sensitive: sa, sb = sa.lower(), sb.lower()
    return sa==sb

def choose_best_key(dfA, dfB):
    common = [c for c in dfA.columns if c in dfB.columns]
    if not common: raise ValueError('Aucune colonne commune entre les deux fichiers.')
    best, best_overlap, best_minfilled = None, -1, -1
    for c in common:
        sA, sB = dfA[c], dfB[c]
        nonA = sA.notna() & (sA.astype(str).str.strip()!='')
        nonB = sB.notna() & (sB.astype(str).str.strip()!='')
        countA, countB = int(nonA.sum()), int(nonB.sum())
        setA = set(normalize_series(sA,True,False)[nonA].tolist())
        setB = set(normalize_series(sB,True,False)[nonB].tolist())
        overlap, minfilled = len(setA & setB), min(countA, countB)
        if overlap>best_overlap or (overlap==best_overlap and minfilled>best_minfilled):
            best, best_overlap, best_minfilled = c, overlap, minfilled
    return best, normalize_series(dfA[best],True,False), normalize_series(dfB[best],True,False), {}

def list_sheet_names(xls_bytes: bytes):
    with pd.ExcelFile(io.BytesIO(xls_bytes)) as xf: return xf.sheet_names

def read_df_from_bytes(xls_bytes: bytes, sheet_name: Any):
    df = pd.read_excel(io.BytesIO(xls_bytes), sheet_name=sheet_name)
    if isinstance(df, dict): df = df[next(iter(df))]
    return df

def write_sorted_sheet(ws, df, title):
    ws.title = title[:31]; header_font = Font(bold=True)
    for j,col in enumerate(df.columns, start=1):
        c = ws.cell(row=1, column=j, value=str(col)); c.font = header_font
    for i in range(len(df)):
        for j,val in enumerate(df.iloc[i].tolist(), start=1):
            ws.cell(row=i+2, column=j, value=val)
    ws.freeze_panes = 'A2'; ws.auto_filter.ref = ws.dimensions
    for j,col in enumerate(df.columns, start=1):
        try: width = min(max(10, int(df[col].astype(str).map(len).max())+2), 60)
        except: width = 20
        ws.column_dimensions[get_column_letter(j)].width = width

def compare_single_sheet(sheet, bytes_a, bytes_b, numeric_tol, decimal_comma, case_sensitive, strip_ws):
    dfA, dfB = read_df_from_bytes(bytes_a, sheet), read_df_from_bytes(bytes_b, sheet)
    dfA.columns, dfB.columns = normalize_colnames(dfA.columns.tolist()), normalize_colnames(dfB.columns.tolist())
    common = [c for c in dfA.columns if c in dfB.columns]
    if not common: return None
    key, keyA_norm, keyB_norm, _ = choose_best_key(dfA, dfB)
    dfA2, dfB2 = dfA.copy(), dfB.copy()
    dfA2['_key_norm_'], dfB2['_key_norm_'] = keyA_norm, keyB_norm
    dfA2['_orig_pos_'], dfB2['_orig_pos_'] = np.arange(len(dfA2)), np.arange(len(dfB2))
    dfA_sorted = dfA2.sort_values(by=['_key_norm_','_orig_pos_'], kind='mergesort').reset_index(drop=True)
    dfB_sorted = dfB2.sort_values(by=['_key_norm_','_orig_pos_'], kind='mergesort').reset_index(drop=True)
    max_len = max(len(dfA_sorted), len(dfB_sorted))
    dfA_cmp, dfB_cmp = dfA_sorted.reindex(range(max_len)).reset_index(drop=True), dfB_sorted.reindex(range(max_len)).reset_index(drop=True)
    common_no_aux = [c for c in common if c not in ['_key_norm_','_orig_pos_']]
    diff_records, total_cells, diff_cells = [], 0, 0
    def get_val(df,i,col):
        try: return df.at[i,col]
        except: return float('nan')
    for i in range(max_len):
        key_val = get_val(dfA_cmp,i,key)
        if pd.isna(key_val) or str(key_val).strip()=='': key_val = get_val(dfB_cmp,i,key)
        for col in common_no_aux:
            a,b = get_val(dfA_cmp,i,col), get_val(dfB_cmp,i,col)
            if not equal_values(a,b,strip_ws,case_sensitive,numeric_tol,decimal_comma):
                diff_cells += 1
                diff_records.append({'sheet':sheet,'row_index_sorted':i+1,'key_value':to_str(key_val),
                                     'column':col,'value_in_A':a,'value_in_B':b})
            total_cells += 1
    differences_df = pd.DataFrame(diff_records, columns=['sheet','row_index_sorted','key_value','column','value_in_A','value_in_B'])
    summary = {'sheet':sheet,'chosen_key':key,'rows_in_A':int(len(dfA_sorted)),'rows_in_B':int(len(dfB_sorted)),
               'common_columns_compared':', '.join(common_no_aux),'total_cells_compared':int(total_cells),
               'different_cells':int(diff_cells),'different_cells_pct':float((diff_cells/total_cells*100) if total_cells else 0.0)}
    return {'sheet':sheet,'summary':summary,
            'dfA_sorted':dfA_sorted.drop(columns=['_key_norm_','_orig_pos_']),
            'dfB_sorted':dfB_sorted.drop(columns=['_key_norm_','_orig_pos_']),
            'differences':differences_df}

def build_report_for_all_sheets(bytes_a, bytes_b, numeric_tol, decimal_comma, case_sensitive, strip_ws):
    sheets = [s for s in list_sheet_names(bytes_a) if s in set(list_sheet_names(bytes_b))]
    if not sheets: raise ValueError('Aucune feuille commune entre A et B.')
    results = [r for s in sheets if (r:=compare_single_sheet(s,bytes_a,bytes_b,numeric_tol,decimal_comma,case_sensitive,strip_ws))]
    if not results: raise ValueError('Aucune feuille comparable (pas de colonnes communes).')
    summary_df = pd.DataFrame([r['summary'] for r in results])
    differences_df = pd.concat([r['differences'] for r in results], ignore_index=True)
    wb = Workbook(); wb.remove(wb.active)
    red = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    for r in results:
        wsA = wb.create_sheet(title=(r['sheet'][:25] + '_A_sorted')); write_sorted_sheet(wsA, r['dfA_sorted'], wsA.title)
        wsB = wb.create_sheet(title=(r['sheet'][:25] + '_B_sorted')); write_sorted_sheet(wsB, r['dfB_sorted'], wsB.title)
        def header_map(ws):
            m={}
            for c in range(1, ws.max_column+1):
                name = ws.cell(row=1,column=c).value
                if name is None: continue
                k = ' '.join(str(name).strip().split()).lower()
                if k not in m: m[k]=c
            return m
        mapA, mapB = header_map(wsA), header_map(wsB)
        for _,d in r['differences'].iterrows():
            col = ' '.join(str(d['column']).strip().split()).lower()
            row = int(d['row_index_sorted'])
            if col in mapA: wsA.cell(row=1+row, column=mapA[col]).fill = red
            if col in mapB: wsB.cell(row=1+row, column=mapB[col]).fill = red
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp: path = tmp.name
    wb.save(path)
    with pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        summary_df.to_excel(writer, index=False, sheet_name='Summary_by_sheet')
        differences_df.to_excel(writer, index=False, sheet_name='Differences_all')
    with open(path,'rb') as f: return f.read()

st.set_page_config(page_title='Comparateur Excel (multi-feuilles)', layout='wide')
st.title('🔎 Comparateur Excel — multi-feuilles, tri & surlignage après tri')
st.markdown('''Dépose deux classeurs **.xlsx** (A & B). Chaque feuille commune est triée sur la meilleure colonne-clé,
puis comparée cellule par cellule. Deux feuilles **_A_sorted** et **_B_sorted** sont générées avec les **écarts en rouge**,
et des onglets analytiques sont ajoutés.''')

with st.sidebar:
    st.header('⚙️ Options')
    numeric_tol = st.number_input('Tolérance numérique', min_value=0.0, value=0.0, step=0.0001, format='%.6f')
    decimal_comma = st.checkbox('Virgule décimale (européenne)', value=False)
    case_sensitive = st.checkbox('Sensible à la casse', value=False)
    strip_ws = st.checkbox('Nettoyer espaces superflus', value=True)
    output_name = st.text_input('Nom du fichier de sortie', value='Differences.xlsx')

col1, col2 = st.columns(2)
with col1: up_a = st.file_uploader('📄 Classeur A (.xlsx)', type=['xlsx'], accept_multiple_files=False)
with col2: up_b = st.file_uploader('📄 Classeur B (.xlsx)', type=['xlsx'], accept_multiple_files=False)

run = st.button('🚀 Générer le rapport multi-feuilles')
if run:
    if up_a is None or up_b is None:
        st.warning('Merci de charger **A** et **B**.')
    else:
        try:
            out_bytes = build_report_for_all_sheets(up_a.getvalue(), up_b.getvalue(), numeric_tol, decimal_comma, case_sensitive, strip_ws)
            st.success('Rapport généré.')
            st.download_button('💾 Télécharger le rapport Excel', data=out_bytes,
                               file_name=output_name or 'Differences.xlsx',
                               mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        except Exception as e:
            st.error(f'Erreur : {e}'); st.exception(e)
